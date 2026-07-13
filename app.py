"""
OntoLink - Generic Ontology-Driven Data Tagger
Run: python app.py  →  open http://localhost:8000

Users load any OWL/RDFS ontology (TTL). The properties panel, AI suggestions,
serialization, and NL query all adapt to whichever ontology is loaded. No
domain-specific logic anywhere in this file.
"""

import io
import json
import os
import re
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from rdflib import Graph, Literal, Namespace, URIRef
from rdflib.namespace import RDF, XSD

from ontology_loader import OntologyConfig, empty_config, load_ontology
from exporters import EXPORT_FORMATS, export as run_export

load_dotenv()

app = FastAPI(title="OntoLink")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "ontologies"
UPLOAD_DIR.mkdir(exist_ok=True)

# ── Ontology state ───────────────────────────────────────────────────────────
# Held in memory. If a TTL file was previously uploaded, load it on startup.
_ontology_lock = threading.Lock()
_ontology: OntologyConfig = empty_config()
_last_ttl_path: Optional[Path] = None


def _try_load_last() -> None:
    """Load the most recently uploaded ontology, if any."""
    global _ontology, _last_ttl_path
    ttls = sorted(UPLOAD_DIR.glob("*.ttl"), key=lambda p: p.stat().st_mtime, reverse=True)
    if ttls:
        try:
            _ontology = load_ontology(ttls[0])
            _last_ttl_path = ttls[0]
        except Exception:
            _ontology = empty_config()


_try_load_last()


def current_ontology() -> OntologyConfig:
    with _ontology_lock:
        return _ontology


def set_ontology(cfg: OntologyConfig, ttl_path: Optional[Path] = None) -> None:
    global _ontology, _last_ttl_path
    with _ontology_lock:
        _ontology = cfg
        if ttl_path:
            _last_ttl_path = ttl_path


def _range_to_xsd(r: str):
    return {
        "string": XSD.string,
        "integer": XSD.integer,
        "int": XSD.integer,
        "float": XSD.float,
        "date": XSD.date,
        "boolean": XSD.boolean,
    }.get(r, XSD.string)


def _slug(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", str(s).strip())[:60]


# ── Anthropic client ─────────────────────────────────────────────────────────
def get_client(request: Optional[Request] = None):
    import anthropic
    key = ""
    if request is not None:
        key = request.headers.get("X-API-Key", "").strip()
    if not key:
        key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not key:
        raise HTTPException(422, "No API key provided. Add your Anthropic API key in Settings.")
    return anthropic.Anthropic(api_key=key)


# ── RDF graph builder ────────────────────────────────────────────────────────
def build_graph(sheets: List[Dict], column_tags: Dict) -> Graph:
    """
    Turn tagged spreadsheets into an RDF knowledge graph using the currently
    loaded ontology. Each row becomes one or more class instances plus the
    object-property links between them. Instances are deduplicated per workbook
    by their key property.

    Tag payload from the frontend must include:
      - uri:   local name of the data property (e.g. "name")
      - class: local name of the ontology class this property lives on
      - range: xsd type family (string/integer/float/date/boolean)
      - label: display label (unused here, kept for completeness)
    """
    onto = current_ontology()
    HXLO = Namespace(onto.namespace)                    # ontology namespace
    HXLD = Namespace(onto.data_namespace)               # data instances namespace

    g = Graph()
    g.bind("onto", HXLO)
    g.bind("data", HXLD)
    g.bind("xsd", XSD)

    if not onto.entity_model:
        return g

    # entity_cache is keyed per workbook to prevent cross-org merging when codes collide
    entity_cache: Dict = {}

    # Priority order: process more-connected classes first (Distribution before Beneficiary etc.)
    priority_index = {c: i for i, c in enumerate(onto.entity_priority)}

    for sheet in sheets:
        name = sheet.get("name", "")
        data = sheet.get("data", [])
        tags = column_tags.get(name, {})
        if not tags or len(data) < 2:
            continue

        wb_prefix = name.split("__")[0] if "__" in name else _slug(name)
        sid = _slug(name)

        # Group tagged columns by their class
        cols_by_class: Dict[str, List] = {}
        for ci_str, tag in tags.items():
            if not isinstance(tag, dict):
                continue
            cls_name = tag.get("class")
            uri = tag.get("uri")
            if not cls_name or not uri:
                continue
            if cls_name not in onto.classes:
                continue
            cols_by_class.setdefault(cls_name, []).append((int(ci_str), tag))

        if not cols_by_class:
            continue

        present_classes = sorted(cols_by_class.keys(),
                                 key=lambda c: priority_index.get(c, 999))

        for ri, row in enumerate(data[1:], 1):
            if not any(str(c).strip() for c in row):
                continue

            row_nodes: Dict[str, URIRef] = {}

            for cls_name in present_classes:
                cols = cols_by_class[cls_name]
                cls = onto.classes[cls_name]

                # Determine identifier for this entity instance on this row
                key_val = None
                if cls.key_property:
                    for ci, tag in cols:
                        if tag.get("uri") == cls.key_property and ci < len(row):
                            v = str(row[ci]).strip()
                            if v:
                                key_val = v
                                break

                if not key_val:
                    # Fallback: concatenate all non-empty values for this class
                    vals = [str(row[ci]).strip()
                            for ci, _ in cols
                            if ci < len(row) and str(row[ci]).strip()]
                    if not vals:
                        continue
                    key_val = "|".join(vals)

                cache_key = (wb_prefix, cls_name, key_val)
                if cache_key in entity_cache:
                    node = entity_cache[cache_key]
                else:
                    node = HXLD[f"{wb_prefix}_{cls_name}_{_slug(key_val)}"]
                    g.add((node, RDF.type, HXLO[cls_name]))
                    entity_cache[cache_key] = node

                for ci, tag in cols:
                    if ci >= len(row):
                        continue
                    val = row[ci]
                    if val is None or str(val).strip() == "":
                        continue
                    g.add((node,
                           HXLO[tag["uri"]],
                           Literal(str(val), datatype=_range_to_xsd(tag.get("range", "string")))))

                row_nodes[cls_name] = node

            # Auto-link entities that co-occur on this row
            for (subj_cls, obj_cls), prop_uri in onto.link_model.items():
                if subj_cls in row_nodes and obj_cls in row_nodes:
                    g.add((row_nodes[subj_cls], HXLO[prop_uri], row_nodes[obj_cls]))

    return g


# ── API: status ──────────────────────────────────────────────────────────────
@app.get("/api/status")
def status(request: Request):
    onto = current_ontology()
    env_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    header_key = request.headers.get("X-API-Key", "").strip()
    return {
        "status": "ok",
        "api_key_configured": bool(env_key or header_key),
        "ontology_loaded": onto.name != "(none)",
        "ontology_name": onto.name,
        "ontology_class_count": onto.frontend_config.get("class_count", 0),
        "ontology_property_count": onto.frontend_config.get("property_count", 0),
    }


# ── API: get current ontology descriptor for the sidebar ────────────────────
@app.get("/api/ontology")
def get_ontology():
    return current_ontology().frontend_config


# ── API: upload a new ontology ──────────────────────────────────────────────
@app.post("/api/ontology/upload")
async def upload_ontology(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".ttl"):
        raise HTTPException(422, "Please upload a Turtle (.ttl) ontology file.")
    content = await file.read()
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", file.filename)
    dest = UPLOAD_DIR / safe_name
    dest.write_bytes(content)
    try:
        cfg = load_ontology(dest, name=file.filename.rsplit(".", 1)[0])
    except Exception as e:
        dest.unlink(missing_ok=True)
        raise HTTPException(422, f"Could not parse ontology: {e}")
    set_ontology(cfg, dest)
    return {
        "ok": True,
        "name": cfg.name,
        "namespace": cfg.namespace,
        "class_count": cfg.frontend_config["class_count"],
        "property_count": cfg.frontend_config["property_count"],
        "link_count": cfg.frontend_config["link_count"],
    }


# ── API: unload / reset ─────────────────────────────────────────────────────
@app.post("/api/ontology/reset")
def reset_ontology():
    set_ontology(empty_config(), None)
    return {"ok": True}


# ── API: upload spreadsheet ─────────────────────────────────────────────────
@app.post("/api/upload")
async def upload_workbook(file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "Could not parse file. Please upload a valid .xlsx file.")
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
            if len(rows) > 1000:
                break
        sheets.append({"name": name, "data": rows})
    wb.close()
    return {"sheets": sheets, "filename": file.filename}


# ── API: serialize to RDF Turtle ────────────────────────────────────────────
class SerializeReq(BaseModel):
    sheets: List[Dict]
    columnTags: Dict


@app.post("/api/serialize")
def serialize(req: SerializeReq):
    onto = current_ontology()
    if onto.name == "(none)":
        raise HTTPException(422, "No ontology loaded. Upload a TTL file first.")
    g = build_graph(req.sheets, req.columnTags)
    if len(g) == 0:
        raise HTTPException(422, "No tagged data found. Please tag at least one column first.")
    return {"turtle": g.serialize(format="turtle"), "triple_count": len(g)}


# ── API: export tagged Excel ────────────────────────────────────────────────
class ExportReq(BaseModel):
    sheets: List[Dict]
    columnTags: Dict
    active_sheet: int = 0


@app.post("/api/export")
def export_workbook(req: ExportReq):
    if req.active_sheet >= len(req.sheets):
        raise HTTPException(422, "No sheet data.")
    sheet = req.sheets[req.active_sheet]
    name = sheet["name"]
    data = sheet["data"]
    tags = req.columnTags.get(name, {})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    headers = data[0] if data else []
    ws.append(headers)
    # Row 2 = the property URIs the columns were tagged as
    ws.append([tags.get(str(ci), {}).get("uri", "") for ci in range(len(headers))])
    for row in data[1:]:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tagged_{_slug(name)}.xlsx"},
    )


# ── API: retailer feed export (JSON-LD, XML, CSV) ───────────────────────────
@app.get("/api/export-formats")
def export_formats():
    """List all supported feed formats."""
    return {
        "formats": [
            {"key": key, "label": spec["label"], "extension": spec["extension"]}
            for key, spec in EXPORT_FORMATS.items()
        ],
        "classes": [
            cls for cls, defn in current_ontology().classes.items()
            if defn.data_properties
        ],
    }


class FeedExportReq(BaseModel):
    sheets: List[Dict]
    columnTags: Dict
    format: str                              # "json-ld" | "google-shopping-xml" | "csv"
    product_class: str = "Product"           # which class in the ontology represents a product row
    store_title: Optional[str] = None
    store_link: Optional[str] = None


@app.post("/api/export-feed")
def export_feed(req: FeedExportReq):
    onto = current_ontology()
    if onto.name == "(none)":
        raise HTTPException(422, "No ontology loaded. Upload a TTL file first.")
    if req.product_class not in onto.classes:
        raise HTTPException(422,
            f"Class '{req.product_class}' not found in the loaded ontology. "
            f"Choose one of: {sorted(onto.classes.keys())}")

    g = build_graph(req.sheets, req.columnTags)
    if len(g) == 0:
        raise HTTPException(422, "No tagged data found. Please tag columns first.")

    extra: Dict[str, Any] = {}
    if req.format == "google-shopping-xml":
        if req.store_title: extra["store_title"] = req.store_title
        if req.store_link: extra["store_link"] = req.store_link

    try:
        result = run_export(g, req.format, product_class=req.product_class, **extra)
    except ValueError as e:
        raise HTTPException(422, str(e))

    return Response(
        content=result["content"],
        media_type=result["mime"],
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"'
        },
    )


# ── API: AI suggest tags ─────────────────────────────────────────────────────
class ColInfo(BaseModel):
    sheet: str
    col_index: int
    header: str
    samples: List[str]


class SuggestAllReq(BaseModel):
    columns: List[ColInfo]


def _property_catalog_for_llm(onto: OntologyConfig) -> str:
    """Compact text listing every property with its class, label and range —
    everything the LLM needs to pick the right one for a column header."""
    lines: List[str] = []
    for cls_name, cls in onto.classes.items():
        if not cls.data_properties:
            continue
        lines.append(f"Class {cls_name} ({cls.label}):")
        for dp in cls.data_properties:
            lines.append(f"  {dp.uri} | label='{dp.label}' | range={dp.range}"
                         + (f" | {dp.comment[:80]}" if dp.comment else ""))
    return "\n".join(lines)


@app.post("/api/suggest-all")
def suggest_all(req: SuggestAllReq, request: Request):
    if not req.columns:
        return {"suggestions": []}
    onto = current_ontology()
    if onto.name == "(none)":
        raise HTTPException(422, "No ontology loaded. Upload a TTL file first.")

    client = get_client(request)
    catalog = _property_catalog_for_llm(onto)
    col_list = "\n".join(
        f'{i}. Sheet="{c.sheet}" Col={c.col_index} Header="{c.header}" Samples=[{", ".join(c.samples[:4])}]'
        for i, c in enumerate(req.columns)
    )

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        system=f"""You map spreadsheet columns to properties of an OWL ontology.

Ontology: {onto.name}

Available properties:
{catalog}

For every input column, choose the single best matching property from the list above.
Respond with a JSON ARRAY, one object per input column IN ORDER:
{{"col_index": N, "sheet": "...", "uri": "<property local name>", "class": "<class name>", "label": "...", "range": "..."}}
If no property fits a column, return: {{"col_index": N, "sheet": "...", "uri": null}}
No prose, no markdown fences.""",
        messages=[{"role": "user", "content": f"Columns:\n{col_list}"}]
    )
    text = msg.content[0].text.strip()
    try:
        parsed = json.loads(re.sub(r"^```(?:json)?|```$", "", text, flags=re.MULTILINE).strip())
    except Exception:
        m = re.search(r"\[.*\]", text, re.DOTALL)
        parsed = json.loads(m.group()) if m else []
    return {"suggestions": parsed}


# ── API: natural-language SPARQL query ──────────────────────────────────────
class QueryReq(BaseModel):
    sheets: List[Dict]
    columnTags: Dict
    question: str


@app.post("/api/query")
def query(req: QueryReq, request: Request):
    onto = current_ontology()
    if onto.name == "(none)":
        raise HTTPException(422, "No ontology loaded. Upload a TTL file first.")
    client = get_client(request)

    g = build_graph(req.sheets, req.columnTags)
    if len(g) == 0:
        raise HTTPException(422, "No tagged data to query. Please tag columns first.")

    # Build schema — the exact classes and properties actually used in the data
    used: Dict[str, set] = {}
    for sheet_tags in req.columnTags.values():
        for tag in sheet_tags.values():
            if not isinstance(tag, dict) or not tag.get("uri") or not tag.get("class"):
                continue
            cls_name = tag["class"]
            if cls_name in onto.classes:
                used.setdefault(cls_name, set()).add(
                    (tag["uri"], tag.get("range", "string"))
                )

    entity_blocks = []
    for cls_name in sorted(used.keys()):
        block = [f"Entity: onto:{cls_name}"]
        for uri, rng in sorted(used[cls_name]):
            block.append(f"  onto:{uri} (type: {rng})")
        entity_blocks.append("\n".join(block))

    active_links = [
        f"  onto:{s} --onto:{p}--> onto:{o}"
        for (s, o), p in onto.link_model.items()
        if s in used and o in used
    ]

    schema_desc = (
        "Entities and their exact properties in this dataset:\n\n"
        + "\n\n".join(entity_blocks)
        + ("\n\nObject-property links:\n" + "\n".join(active_links) if active_links else "")
    )

    system_prompt = f"""You write SPARQL for an RDF knowledge graph.

Namespace prefix:
  PREFIX onto: <{onto.namespace}>

Rules:
- A property lives on the entity it is listed under in the schema. Never attach it to any other entity.
- Combine information across entities using the object-property links.
- Wrap potentially-missing properties in OPTIONAL.
- Use SELECT DISTINCT.
- Do NOT put entity node variables in SELECT — only the literal value variables the user asked for.
- If an entity may link to multiple values, aggregate with GROUP BY + GROUP_CONCAT.

Return ONLY JSON (no markdown):
{{"sparql":"SELECT ... WHERE {{ ... }}","explanation":"one-sentence description"}}"""

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2000,
        system=system_prompt,
        messages=[{"role": "user", "content": f"{schema_desc}\n\nQuestion: {req.question}"}]
    )
    text = msg.content[0].text.strip()

    ai: Dict = {}
    clean = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip(), flags=re.MULTILINE).strip()
    try:
        ai = json.loads(clean)
    except Exception:
        # Balanced-brace scan (handles nested { } inside SPARQL string)
        start = clean.find("{")
        if start >= 0:
            depth = 0
            in_str = False
            esc = False
            for i in range(start, len(clean)):
                c = clean[i]
                if esc:
                    esc = False
                    continue
                if c == "\\":
                    esc = True
                    continue
                if c == '"':
                    in_str = not in_str
                    continue
                if in_str:
                    continue
                if c == "{":
                    depth += 1
                elif c == "}":
                    depth -= 1
                    if depth == 0:
                        try:
                            ai = json.loads(clean[start:i + 1])
                        except Exception:
                            pass
                        break

    sparql = ai.get("sparql", "")
    if not sparql:
        preview = text[:300].replace("\n", " ")
        return {"sparql": "", "columns": [], "rows": [],
                "summary": f"Could not parse a SPARQL query from the AI response. Try rephrasing. Raw: {preview}",
                "triple_count": len(g)}

    try:
        results = g.query(sparql)
        columns = [str(v) for v in (results.vars or [])]
        rows = [[str(r[v]) if r[v] is not None else "" for v in results.vars] for r in results]

        # Clean up: drop columns that only contain internal node URIs
        if rows:
            data_prefix = onto.data_namespace
            keep = [i for i in range(len(columns))
                    if not all(v.startswith(data_prefix) for v in (r[i] for r in rows) if v)]
            if keep and len(keep) < len(columns):
                columns = [columns[i] for i in keep]
                rows = [[r[i] for i in keep] for r in rows]

        # Deduplicate exact rows
        seen, unique = set(), []
        for r in rows:
            t = tuple(r)
            if t not in seen:
                seen.add(t)
                unique.append(r)
        rows = unique

        # Remove subsumed rows (rows that are identical to another except for blanks)
        if len(rows) <= 3000:
            def subsumed(a, b):
                return a != b and all(x == "" or x == y for x, y in zip(a, b))
            rows = [r for r in rows if not any(subsumed(r, o) for o in rows)]
    except Exception as e:
        return {"sparql": sparql, "columns": [], "rows": [],
                "summary": f"SPARQL error: {e}", "triple_count": len(g)}

    summary = (ai.get("explanation", "") + f" Found {len(rows)} result(s).").strip()
    return {"sparql": sparql, "columns": columns, "rows": rows,
            "summary": summary, "triple_count": len(g)}


# ── API: legacy set-key (kept for compatibility) ────────────────────────────
class SetKeyReq(BaseModel):
    key: str


@app.post("/api/set-key")
def set_key(req: SetKeyReq):
    os.environ["ANTHROPIC_API_KEY"] = req.key.strip()
    return {"ok": True}


# ── Frontend ────────────────────────────────────────────────────────────────
frontend_dir = BASE_DIR / "frontend"
if frontend_dir.exists():
    app.mount("/", StaticFiles(directory=str(frontend_dir), html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    print("\n" + "=" * 50)
    print("  OntoLink - Generic Ontology-Driven Data Tagger")
    print("  Open: http://localhost:8000")
    print("=" * 50 + "\n")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=False)
